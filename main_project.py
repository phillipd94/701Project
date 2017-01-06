# -*- coding: utf-8 -*-
"""
Created on Tue Nov 22 13:43:17 2016

@author: Phillip Dix
"""

#should probably figure out how unit tests work and add some to this thing
#should probably put some docstrings and stuff in here too

from __future__ import division
import traceback
import sys
import numpy as np
import os
import re
import openpyxl as xl
import encodeindex
from codeParser import codeParser
import pickle

from PyQt5.QtWidgets import (QTabWidget,QApplication, QDialog, QLineEdit, QVBoxLayout,QHBoxLayout, QGridLayout, QTableWidget, QTableWidgetItem,
                             QMainWindow, QAction, QFileDialog, QMessageBox, QComboBox, QTextEdit, QPushButton,QHeaderView)
import PyQt5.QtGui as QtGui
from PyQt5.QtCore import QObject, pyqtSignal, Qt,QRegExp

from PyQt5.Qt import QFrame, QWidget, QPainter
 

class EmittingStream(QObject): #http://stackoverflow.com/questions/8356336/how-to-capture-output-of-pythons-interpreter-and-show-in-a-text-widget

    textWritten = pyqtSignal(str)

    def write(self, text):
        try:
            self.textWritten.emit(str(text))
        except:
            msg=QMessageBox()
            msg.setText('error: Phillip is bad at code')
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

    


class MainWindow(QMainWindow):
    
    def __init__(self, parent=None) :
        super(MainWindow, self).__init__(parent)
   
        # Create the file menu
        
        self.menuFile = self.menuBar().addMenu("&File")
        self.actionSaveAs = QAction("&Save Spreadsheet As", self)
        self.actionSaveAs.triggered.connect(self.saveas)
        
        self.actionSSaveAs = QAction("&Save Script As", self)
        self.actionSSaveAs.triggered.connect(self.ssaveas)
        
        self.actionclosetab = QAction("&Close Current WorkBook", self)
        self.actionclosetab.triggered.connect(self.closewb)
        
        self.actionSOpen = QAction("&Open Script", self)
        self.actionSOpen.triggered.connect(self.openscript)
        ########SAVE THIS IT'S CLOSE TO WORKING
#        self.actionVSaveAs = QAction("&Save Variable Space As", self)
#        self.actionVSaveAs.triggered.connect(self.vsaveas)
#        
#        self.actionOpenvars = QAction("&Open Variable Space", self)
#        self.actionOpenvars.triggered.connect(self.openvars)
        #############################################
        self.actionNewSpreadsheet = QAction("&New Spreadsheet", self)
        self.actionNewSpreadsheet.triggered.connect(self.newspreadsheet)
        
        self.actionOpenSpreadsheet = QAction("&Open Spreadsheet", self)
        self.actionOpenSpreadsheet.triggered.connect(self.openspreadsheet)
        
        self.actionQuit = QAction("&Quit", self)
        self.actionQuit.triggered.connect(self.close)#self.close is predefined, so I didn't have to change anything here
        self.menuFile.addActions([self.actionNewSpreadsheet,self.actionOpenSpreadsheet,self.actionSaveAs,self.actionSSaveAs,self.actionSOpen,self.actionclosetab, self.actionQuit])
        
        # Set the central widget
        self.widget1 = Form()
        self.setCentralWidget(self.widget1)
        
        sys.stdout = EmittingStream(textWritten=self.normalOutputWritten)
        
        

#        sys.stderr = EmittingStream(textWritten=self.normalOutputWritten) ###UNCOMMENT THIS OF YOU WANT INTERNAL ERRORS REROUTED TO THE OUTPUT BOX
        self.newspreadsheet()


########SAVE THIS IT'S CLOSE TO WORKING
#    def vsaveas(self):
#        fname = unicode(QFileDialog.getSaveFileName(self, "Save Variable Space as...")[0])
#        if fname:
#            b=open(fname,'w')
#            pickle.dump(self.widget1.locals,b)
#            b.close()
#
#    def openvars(self):
#        fname = unicode(QFileDialog.getOpenFileName(self, "Open")[0])
#        try:
#            b=open(fname,'r')
#            pickle.load(self.widget1.locals,b)
#            b.close()
#        except:
#            print "error, bad file type"
#################
    def closewb(self):
        del self.widget1.scr.edit.activeWB[self.widget1.scr.edit.WBindex]
        del self.widget1.scr.edit.activeWS[self.widget1.scr.edit.WBindex]
        self.widget1.table.removeTab(self.widget1.scr.edit.WBindex)

    def ssaveas(self):
        fname = unicode(QFileDialog.getSaveFileName(self, "Save Script as...")[0])  #AT LAST I HAVE MY VICTORY, THIS METHOD RETURNS A TUPLE, AND I HAD TO SELECT A SPECIFIC VALUE TO PASS TO THE UNICODE TYPE CAST
        if fname :
            try:
                 a=self.widget1.scr.edit.toPlainText()
                 b=open(fname+".py","w")
                 b.write(a)
                 b.close
            except:
                print "error saving file"  
                
    def openscript(self):
        fname = unicode(QFileDialog.getOpenFileName(self, "Open")[0])
        if '.py' in fname or '.txt' in fname:
            a=open(fname,'r')
            b=a.read()
            self.widget1.scr.edit.setPlainText(b)
        
    def newspreadsheet(self):

        self.widget1.scr.edit.activeWB.append(xl.Workbook())
        d=self.widget1.scr.edit.activeWB[len(self.widget1.scr.edit.activeWB)-1].worksheets
        d[0]["A1"]=0
        self.widget1.scr.edit.activeWS.append(d)
        z=QTableWidget(0,0,self)
        v=QTabWidget()
        v.currentChanged.connect(self.widget1.switch_tabs_ws)
        s=self.widget1.scr.edit.activeWS[0][0]
        for i in s['A1':'Z100']:
                for j in i:
                    j.value=''
        v.addTab(z,"NewSheet")
        self.widget1.table.addTab(v,"Book1")
        self.widget1.updateUI()
        
        
    def openspreadsheet(self):
        fname = unicode(QFileDialog.getOpenFileName(self, "Open")[0])
        if '.xlsx' in fname:
            self.widget1.scr.edit.WBNames.append(fname)
            self.widget1.scr.edit.activeWB.append(xl.load_workbook(fname))
            d=self.widget1.scr.edit.activeWB[len(self.widget1.scr.edit.activeWB)-1].worksheets
            self.widget1.scr.edit.activeWS.append(d)
            v=QTabWidget()
            for m in range(0,len(d)):
                z=QTableWidget(0,0,self)
                z.setRowCount(d[m].max_row)
                z.setColumnCount(d[m].max_column)
                a=d[m].iter_rows()
                for i in range(0,int(d[m].max_row)):
                    try:
                        b=a.next()
                        for j in range(0,int(d[m].max_column)):
                            if (str(b[j].value) != 'None'):
                                z.setItem(i,j,QTableWidgetItem(str(b[j].value)))
                    except:
                        pass
                v.addTab(z,str(d[m].title))
                v.currentChanged.connect(self.widget1.switch_tabs_ws)
            self.widget1.table.addTab(v,str(fname))
        else:
            msg=QMessageBox()
            msg.setText('error: bad file type')
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
        
    def saveas(self) :
        """ Save the computed data as a text file. """


        fname = unicode(QFileDialog.getSaveFileName(self, "Save as...")[0])
        if fname :
            try:
                 self.widget1.scr.edit.activeWB[self.widget1.scr.edit.WBindex].save(filename = fname)
            except:
                 print "error saving file"
                
    def normalOutputWritten(self, text):
        """Append text to the QTextEdit."""

        if text!="\n":
            self.widget1.cmd.append("\n"+text)
                
class Form(QDialog) :

    def __init__(self, parent=None) :
        super(Form, self).__init__(parent)

        # Define three text boxes.
        self.scr=LineTextWidget()
        self.scr.edit.setPlainText("for i in range(0,5):\n\ta='output_test_'+str(i)\n\tb=open(a,'w')\n\tb.write(a)\n\tb.close\n")
        self.run_button = QPushButton('Run', self)
        self.clear_button = QPushButton('Clear Variables',self)
        self.clear_button.clicked.connect(self.clear_vars)
        self.cmd=outputbox()
        
        self.command=commandLine()
        self.command.returnPressed.connect(self.userExec) 

        self.table = QTabWidget(self)
        
        self.vars = QTableWidget(5,2,self)
        self.vars.setColumnCount(2)

        header = self.vars.horizontalHeader()
        header.setSectionResizeMode(3)
        header.setStretchLastSection(True)

        # Define the layout
        layout = QGridLayout()
        layout1=QVBoxLayout()
        layout1.addWidget(self.run_button)
        layout1.addWidget(self.scr)
        layout1.addWidget(self.command)
        layout1.addWidget(self.cmd)
        layout2=QHBoxLayout()
        layout2.addLayout(layout1)
        layout3=QVBoxLayout()
        layout3.addWidget(self.clear_button)
        layout3.addWidget(self.vars)
        layout2.addLayout(layout3)
        layout.addWidget(self.table,0,0)
        layout.addLayout(layout2,0,1)
        self.setLayout(layout)
        self.table.currentChanged.connect(self.switch_tabs_wb)
		
        # Connect our output box to the update function
        try:
            self.run_button.clicked.connect(self.userExec) 
        except:
            print "error"
        self.locals={}
        self.flag=True
		

    def clear_vars(self):
        
        self.locals=()
        self.flag=True
        self.vars.setColumnCount(2)
        self.vars.setRowCount(5)
        for i in range(0,self.vars.columnCount()):
            for j in range(0,self.vars.rowCount()):
                self.vars.setItem(i,j,QTableWidgetItem(""))

        
    def userExec(self) :
        """ Method for executing user code"""
        signal_sender=self.sender()

        flag_error_exception=True
        try:
            if self.flag:#is this the first execution? if so, no need to pass in self.locals because it doesn't exist yet
                if isinstance(signal_sender, QPushButton):#check the sender type so we know where to get the code from
                    exec(self.scr.edit.getCode())#execute code from script box
                    self.flag=False
                elif isinstance(signal_sender, commandLine):
                    exec(self.command.getCode())#execute code from command line
                    self.flag=False
                else:
                    print "error in execution handling:1"
            else:#if this is not the first execution of user code, self.locals is populated (unless the user pressed clear vars) and needs to be passed into exec
                if isinstance(signal_sender, QPushButton):
                    glob=globals()
                    exec(self.scr.edit.getCode(),glob,self.locals)
                elif isinstance(signal_sender, commandLine):
                    glob=globals()
                    exec(self.command.getCode(),glob,self.locals)
                else:
                    print "error in execution handling:2"
        except SyntaxError as err:  #http://stackoverflow.com/questions/28836078/how-to-get-the-line-number-of-an-error-from-exec-or-execfile-in-python
            flag_error_exception=False
            error_class = err.__class__.__name__
            detail = err.args[0]
            line_number = err.lineno
        except Exception as err:
            flag_error_exception=False
            error_class = err.__class__.__name__
            detail = err.args[0]
            cl, exc, tb = sys.exc_info()
            line_number = traceback.extract_tb(tb)[-1][1]
        else:
            pass
        if not flag_error_exception:
            if isinstance(signal_sender, QPushButton):
                print("%s at line %d: %s" % (error_class, line_number-2, detail))#print error details
            else:
                print("%s: %s" % (error_class, detail))#don't give line numbers if it was a command line execution, because that would make no sense


        self.updateUI()  #update the spreadsheet viewer
        

	#set variables in self.locals to show up in the variable viewer, with exception of those declared in the exec function above
        count=0
        self.vars.setRowCount(len(self.locals)-4)
        for i in self.locals:
            if (i!="wb") and (i!="flag_error_exception") and (i!="self") and (i!="signal_sender"):
                self.vars.setItem(count,0,QTableWidgetItem(str(i)))
                self.vars.setItem(count,1,QTableWidgetItem(str(self.locals[i])))
                count+=1

    def addWS(self,a):
        self.scr.edit.activeWB[self.scr.edit.WBindex].create_sheet(a)
        self.scr.edit.activeWS[self.scr.edit.WBindex].append(self.scr.edit.activeWB[self.scr.edit.WBindex].worksheets[-1])
        return self.scr.edit.activeWS
        
    def addWB(self,a):

        self.scr.edit.activeWB.append(xl.Workbook())
        d=self.scr.edit.activeWB[-1].worksheets
        self.scr.edit.activeWS.append(d)
        z=QTableWidget(0,0,self)
        v=QTabWidget()
        v.currentChanged.connect(self.switch_tabs_ws)
        s=self.scr.edit.activeWS[-1][0]
        for i in s['A1':'Z100']:
                for j in i:
                    j.value=''
        v.addTab(z,a)
        self.table.addTab(v,a)
        return self.scr.edit.activeWS
        
        
        
    def updateUI(self):
        k=self.scr.edit.activeWB        
        for s in range(0,len(k)):
     
            d=self.scr.edit.activeWB[s].worksheets
            for i in range(0,len(d)-self.table.widget(s).count()):
                c=self.table.widget(s).count()
                z=QTableWidget(1,1,self)
                if(d[c]["A1"]==None):
                    d[c]["A1"]=0
                self.scr.edit.activeWS[s].append(d[c])
                self.table.widget(s).addTab(z,str(d[c].title))
            for v in range(0,len(d)):
                z=self.table.widget(s).widget(v)
                z.setRowCount(d[v].max_row)
                z.setColumnCount(d[v].max_column)
                a=d[v].iter_rows()
                for i in range(0,int(d[v].max_row)):
                        b=a.next()

                        for j in range(0,int(d[v].max_column)):
                            if (str(b[j].value) != 'None'):
                                z.setItem(i,j,QTableWidgetItem(str(b[j].value)))
                for i in range(0,int(d[v].max_column)):
                    z.setHorizontalHeaderItem(i,QTableWidgetItem(encodeindex.encodeIndex(i+1)))#account for the nonzero indexing of columns in excel

    def switch_tabs_ws(self):
        try:
            self.scr.edit.WSindex=self.table.widget(self.scr.edit.WBindex).currentIndex()
        except AttributeError: #this is called when the tab is first initiallized but before there's a workbook and it causes errors
            pass

    def switch_tabs_wb(self):
        self.scr.edit.WBindex=self.table.currentIndex()
        self.switch_tabs_ws()
#        print self.scr.edit.WBindex
        
class script_box(QTextEdit):

    def __init__(self, parent=None) :
        super(script_box, self).__init__(parent) 
#        self.setAcceptRichText(False)
        self.activeWB=[]
        self.activeWS=[]
        self.WBindex=0
        self.WSindex=0
        self.WBNames=[]
        self.setTabStopWidth(50)
        self.textChanged.connect(self.updateScript)
        self.parseCode=codeParser()

    def updateScript(self):
        
        tempcurs=self.textCursor()
        a=tempcurs.position()
        script=self.document().toHtml()
        script=self.parseCode.getInput(script)
        
        self.blockSignals(True) #block signals to prevent infinite loops
        self.document().setHtml(script)
        self.blockSignals(False)
        tempcurs.setPosition(a)
        self.setTextCursor(tempcurs)

    def getCode(self):
        #method for parsing and returning user input code
        script=self.document().toPlainText()
        script=self.parseCode.getCode(script)
        return script
   

        #####THIS MIGHT BE UNNECCESSARY
    def getInput(self):
        #method for parsing and colorizing keywords for cell names
        script=self.document().toHtml()
        script=self.parseCode.getInput(script)
        return script
        #################
        
    def keyPressEvent(self, event):
        #this adds automatic insertion of tabs where appropriate
        key = event.key()
        super(script_box,self).keyPressEvent(event)
        print self.document().toPlainText().split("\n")
        print key
        if key == 16777220:#found this through trial and error, not even sure if it'll work on every computer
            a=self.document().toPlainText().split("\n")[-2]
            for i in range(0,a.count("\t")):
                self.insertPlainText("\t")
#                print "AAA"
            if ":" in a:
                self.insertPlainText("\t")
#                print "BBB"
#            print a

        


"""NOT MY CODE"""
#code from https://john.nachtimwald.com/2009/08/15/qtextedit-with-line-numbers/
#MIT license


#I updated this to work with pyqt5 and added my own subclasses into it

class LineTextWidget(QFrame):
 
    class NumberBar(QWidget):
 
        def __init__(self, *args):
            QWidget.__init__(self, *args)
            self.edit = None
            # This is used to update the width of the control.
            # It is the highest line that is currently visibile.
            self.highest_line = 0
 
        def setTextEdit(self, edit):
            self.edit = edit
 
        def update(self, *args):
            '''
            Updates the number bar to display the current set of numbers.
            Also, adjusts the width of the number bar if necessary.
            '''
            # The + 4 is used to compensate for the current line being bold.
            width = self.fontMetrics().width(str(self.highest_line)) + 4
            if self.width() != width:
                self.setFixedWidth(width)
            QWidget.update(self, *args)
 
        def paintEvent(self, event):
            contents_y = self.edit.verticalScrollBar().value()
            page_bottom = contents_y + self.edit.viewport().height()
            font_metrics = self.fontMetrics()
            current_block = self.edit.document().findBlock(self.edit.textCursor().position())
 
            painter = QPainter(self)
 
            line_count = 0
            # Iterate over all text blocks in the document.
            block = self.edit.document().begin()
            while block.isValid():
                line_count += 1
 
                # The top left position of the block in the document
                position = self.edit.document().documentLayout().blockBoundingRect(block).topLeft()
 
                # Check if the position of the block is out side of the visible
                # area.
                if position.y() > page_bottom:
                    break
 
                # We want the line number for the selected line to be bold.
                bold = False
                if block == current_block:
                    bold = True
                    font = painter.font()
                    font.setBold(True)
                    painter.setFont(font)
 
                # Draw the line number right justified at the y position of the
                # line. 3 is a magic padding number. drawText(x, y, text).
                painter.drawText(self.width() - font_metrics.width(str(line_count)) - 3, round(position.y()) - contents_y + font_metrics.ascent(), str(line_count))
 
                # Remove the bold style if it was set previously.
                if bold:
                    font = painter.font()
                    font.setBold(False)
                    painter.setFont(font)
 
                block = block.next()
 
            self.highest_line = line_count
            painter.end()
 
            QWidget.paintEvent(self, event)
 
 
    def __init__(self, *args):
        QFrame.__init__(self, *args)
 
        self.setFrameStyle(QFrame.StyledPanel | QFrame.Sunken)
 
        self.edit = script_box()
        self.edit.setFrameStyle(QFrame.NoFrame)
        
 
        self.number_bar = self.NumberBar()
        self.number_bar.setTextEdit(self.edit)
 
        hbox = QHBoxLayout(self)
        hbox.setSpacing(0)
#        hbox.setMargin(0)
        hbox.addWidget(self.number_bar)
        hbox.addWidget(self.edit)
 
        self.edit.installEventFilter(self)
        self.edit.viewport().installEventFilter(self)
 
    def eventFilter(self, object, event):
        # Update the line numbers for all events on the text edit and the viewport.
        # This is easier than connecting all necessary singals.
        if object in (self.edit, self.edit.viewport()):
            self.number_bar.update()
            return False
        return QFrame.eventFilter(object, event)
 
    def getTextEdit(self):
        return self.edit
"""MY CODE AGAIN"""


class outputbox(QTextEdit):
    
    def __init__(self, parent=None) :
        super(outputbox, self).__init__(parent) 
        self.setReadOnly(True)
        self.textChanged.connect(self.restrict_len)
        
    def restrict_len(self):
        if self.document().lineCount()>200:
            self.blockSignals(True) #block signals to prevent infinite loops
            a=self.toPlainText()
            a="\n".join(a.split("\n")[(self.document().lineCount()-190):])
            self.setPlainText(a)
            self.blockSignals(False)


class commandLine(QLineEdit):
    
    def __init__(self, parent=None) :
        super(commandLine, self).__init__(parent) 
        self.setFrame=False
        self.pastcommands=[]
        self.pastcommandsindex=-1
        self.parseCode=codeParser()



    def keyPressEvent(self, event):
        key = event.key()
        if key == Qt.Key_Up:
            
            if self.pastcommandsindex >= len(self.pastcommands)-1:
                pass
            else:
                self.pastcommandsindex+=1
                self.setText(self.pastcommands[self.pastcommandsindex])
                
        if key == Qt.Key_Down:
            if self.pastcommandsindex <= 0:
                pass
            elif self.pastcommandsindex == -1:
                pass
            else:
                self.pastcommandsindex-=1
                self.setText(self.pastcommands[self.pastcommandsindex])
        super(commandLine,self).keyPressEvent(event)
                             

    def getCode(self):
        #method for parsing and returning user input code
        script=self.text()
        
        self.pastcommands.insert(0,script)
        if len(self.pastcommands)>20:
            del self.pastcommands[-1]
        self.pastcommandsindex=-1

        script=self.parseCode.getCode(script)
        self.setText("")
        return script
        
        
#if __name__ == "__main__" :
app = QApplication(sys.argv)
main = MainWindow()
main.show()
app.exec_()
